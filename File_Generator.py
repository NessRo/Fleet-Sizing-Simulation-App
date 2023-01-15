from multiprocessing.connection import wait
from time import sleep
from openpyxl import Workbook
import matplotlib.pyplot as plt
import matplotlib.ticker as mtick
import openpyxl
import os, uuid
from azure.storage.blob import BlobServiceClient, BlobClient, ContainerClient, __version__


connect_str = ''

def Make_file(results, process_id, fp):
    
    wb = Workbook()
    ws = wb.active

    ws.append(results['Fleet Size'])
    ws.append(results['Expected Yearly Demand'])
    ws.append(results['Actuals'])
    ws.append(results['Service Level'])
    ws.insert_cols(1,1)
    ws['A1'] = 'Fleet Size'
    ws['A2'] = 'Expected Yearly Demand'
    ws['A3'] = 'Actuals'
    ws['A4'] = 'Service Level'

    fig, ax = plt.subplots(constrained_layout=True)
    ax_2 = ax.twinx()
    ax_2.set_ylim([0.80,1.00])
    ax_2.yaxis.set_major_formatter(mtick.PercentFormatter(1))

    ax.plot(results["Fleet Size"], results["Expected Yearly Demand"], 'r', label= "Expected yearly Demand")
    ax.plot(results["Fleet Size"], results["Actuals"], 'b', label= "Actuals")
    ax_2.plot(results["Fleet Size"], results["Service Level"], 'g',label= "Service Level")

    ax.set_xlabel("Fleet Size")
    ax.set_ylabel("Cars Released - in a year")
    ax_2.set_ylabel("Service Level")


    handles, labels = [(a + b) for a, b in zip(ax.get_legend_handles_labels(), ax_2.get_legend_handles_labels())]
    plt.legend(handles, labels, loc='lower right')

    plt.savefig("plot.png", dpi=80)

    img = openpyxl.drawing.image.Image("plot.png")
    img.anchor = 'B6'
    ws.add_image(img)

    wb.save(fp)

    dest_file = process_id + '.xlsx'

    blob_service_client = BlobServiceClient.from_connection_string(connect_str)

    blob_client = blob_service_client.get_blob_client(container='sim-results', blob=dest_file)

    with open(fp, "rb") as data:
        blob_client.upload_blob(data)


def make_test_file(fp,dest_file):


    blob_service_client = BlobServiceClient.from_connection_string(connect_str)

    file = open(fp, 'w')
    file.write("Hello, World!")
    file.close()

    blob_client = blob_service_client.get_blob_client(container='sim-results', blob=dest_file)

    with open(fp, "rb") as data:
        blob_client.upload_blob(data)

def Make_file_manual(results, fp):

    wb = Workbook()
    ws = wb.active

    ws.append(results['Fleet Size'])
    ws.append(results['Expected Yearly Demand'])
    ws.append(results['Actuals'])
    ws.append(results['Service Level'])
    ws.insert_cols(1,1)
    ws['A1'] = 'Fleet Size'
    ws['A2'] = 'Expected Yearly Demand'
    ws['A3'] = 'Actuals'
    ws['A4'] = 'Service Level'

    fig, ax = plt.subplots(constrained_layout=True)
    ax_2 = ax.twinx()
    ax_2.set_ylim([0.80,1.00])
    ax_2.yaxis.set_major_formatter(mtick.PercentFormatter(1))

    ax.plot(results["Fleet Size"], results["Expected Yearly Demand"], 'r', label= "Expected yearly Demand")
    ax.plot(results["Fleet Size"], results["Actuals"], 'b', label= "Actuals")
    ax_2.plot(results["Fleet Size"], results["Service Level"], 'g',label= "Service Level")

    ax.set_xlabel("Fleet Size")
    ax.set_ylabel("Cars Released - in a year")
    ax_2.set_ylabel("Service Level")


    handles, labels = [(a + b) for a, b in zip(ax.get_legend_handles_labels(), ax_2.get_legend_handles_labels())]
    plt.legend(handles, labels, loc='lower right')

    plt.savefig("plot.png", dpi=80)

    img = openpyxl.drawing.image.Image("plot.png")
    img.anchor = 'B6'
    ws.add_image(img)

    wb.save(fp)

